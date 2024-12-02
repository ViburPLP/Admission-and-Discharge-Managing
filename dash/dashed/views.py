from datetime import datetime, date
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.db.models import Q, F, Sum, Count
from django.http import HttpResponse, HttpResponseRedirect, FileResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.timezone import is_naive, make_naive, now
from weasyprint import HTML
import csv
import io
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.express as px
import tempfile
from django.db.models.functions import Lower, Trim

from .forms import UpdateForm, UserDetailForm, SchemeForm, ProviderForm
from .models import (
    Member_Detail,
    InsuranceDetail,
    Scheme,
    Provider,
    Admission_details,
    Discharge_details,
    Daily_update,
    Case_manager_update,
)

 #**************************************************************************************************************************************************
 #Admissions Management
@login_required
def home(request):
    updates = Case_manager_update.objects.all().order_by('-date')
    current_date = now().date()

    if request.method == 'POST':
        if 'add_update' in request.POST:
            update= request.POST.get('update')

            if update:
                new_update = Case_manager_update.objects.create(
                    by=request.user,  
                    date=now(),
                    update=update
                )
                new_update.save()
                messages.success(request, 'Update added successfully!')
            else:
                messages.error(request, 'Both date and update are required.')

            return redirect('home') 
    return render(request, 'template/index.html' , {'updates': updates, 'current_date': current_date})
@login_required
def pending_admissions(request): #list of recently imported, awaiting admission.
    query = request.GET.get('q')
    payer_filter = request.GET.get('payer')
    scheme_filter = request.GET.get('scheme')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    order = request.GET.get('order', 'added_at')

    pending_members = Member_Detail.objects.filter(admission_status= 'pending')

    if query:
        pending_members = pending_members.filter(
            Q(name__icontains=query) |
            Q(membership_number__icontains=query)
        )
    if payer_filter:
        pending_members = pending_members.filter(payer__icontains=payer_filter)
    if scheme_filter:
        pending_members = pending_members.filter(scheme__icontains=scheme_filter)
    if date_from:
        pending_members = pending_members.filter(added_at__gte=parse_date(date_from))
    if date_to:
        pending_members = pending_members.filter(added_at__lte=parse_date(date_to))

    pending_members = pending_members.order_by(order)
    
    return render(request, 'template/admissions/pending.html', 
                  {'pending_members': pending_members, 
                    'query': query, 'payer_filter': payer_filter, 
                    'scheme_filter': scheme_filter, 
                    'date_from': date_from, 'date_to': date_to, 
                    'order': order})


@login_required
def admitting_member_detail(request, pk): #page displaying details of the member to be admited.
    member = get_object_or_404(Member_Detail, pk=pk)
    insurance_details = InsuranceDetail.objects.filter(member=member)
    previous_admissions = Discharge_details.objects.filter(
        name=member.name,
        membership_number=member.membership_number
        )

    try:
        scheme = Scheme.objects.get(name=member.scheme)
        providers = scheme.providers.all()
    except Scheme.DoesNotExist: # if the scheme is not registered then shown an empty list of the providers.
        scheme = None
        providers = []

    return render(request, 'template/admissions/admission-page.html', 
                  {'member': member, 
                   'insurance_details': insurance_details,
                   'scheme': scheme,
                   'providers': providers,                 
                    'previous_admissions': previous_admissions,
                    'current_date': timezone.now().date(),
                   })

@login_required
def admit_member(request, pk): # Updated The button to admit a member
    member = get_object_or_404(Member_Detail, pk=pk)
    providers = Provider.objects.all()
    insurance_details = member.insurance_details.all()

    if request.method == 'POST':
        service_provider_id = request.POST.get('service_provider')
        admission_date = request.POST.get('admission_date')
        admission_diagnosis = request.POST.get('admission_diagnosis_description')
        cover_used_id = request.POST.get('cover_used')
        initial_cover_value = request.POST.get('initial_cover_value')
        initial_cover_balance = request.POST.get('initial_cover_balance')
        requested_amount = request.POST.get('requested_amount')
        lou_issued = request.POST.get('lou_amount')

        # Get the selected provider and cover used
        provider = get_object_or_404(Provider, pk=service_provider_id)
        cover_used = get_object_or_404(insurance_details, pk=cover_used_id)

        # Save the admission details to a different table. 
        admission_detail = Admission_details(
            member=member,
            Provider=provider,
            admission_date=admission_date,
            admission_diagnosis=admission_diagnosis,
            cover_used=cover_used.cover_type,
            initial_cover_balance=initial_cover_balance,
            requested_amount=requested_amount,
            lou_issued=lou_issued,
            admited_by= f"{request.user.first_name} {request.user.last_name}"  #registering the authenticated user
        )
        member.admission_status = 'admitted'
        member.save()
        admission_detail.save()

        context = {
            'member': member,
            'admission_details': admission_detail,
            'current_date': timezone.now().date(),
        }

        # Creating the admission LOU
        html_string = render_to_string('template/admissions/admission_lou.html', context)
        pdf_io = io.BytesIO()
        HTML(string=html_string).write_pdf(target=pdf_io)
        pdf_io.seek(0)

        # Create a FileResponse for downloading the PDF
        response = FileResponse(pdf_io, as_attachment=True, filename=f'Admission LOU - {member.name}.pdf')
        return response
    
    return render(request, 'pending_admissions/admitting_member_detail.html', {
        'member': member,
        'insurance_details': insurance_details,
    })


@login_required
def current_admissions(request): #active admissions list. 
    query = request.GET.get('q')
    payer_filter = request.GET.get('payer')
    scheme_filter = request.GET.get('scheme')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    order = request.GET.get('order', 'added_at')

    admitted_members = Member_Detail.objects.filter(admission_status= 'admitted') #filtering the admitted members


    if query: #filtering by name or membership number
        admitted_members = admitted_members.filter(
            Q(name__icontains=query) |
            Q(membership_number__icontains=query)
        )
    if payer_filter: #filtering by payer
        admitted_members = admitted_members.filter(payer__icontains=payer_filter)

    if scheme_filter: #filtering by scheme
        admitted_members = admitted_members.filter(scheme__icontains=scheme_filter)

    if date_from:  #filtering by date
        admitted_members = admitted_members.filter(admission_details__admission_date__gte=parse_date(date_from))

    if date_to: #filtering by date
        admitted_members = admitted_members.filter(admission_details__admission_date__lte=parse_date(date_to))

    admitted_members = admitted_members.order_by(order) #sorting by date
    
    return render(request, 'template/admissions/admitted.html', 
                  {'admitted_members': admitted_members, 
                    'query': query, 'payer_filter': payer_filter, 
                    'scheme_filter': scheme_filter, 
                    'date_from': date_from, 'date_to': date_to, 
                    'order': order})

@login_required
def discharging_member_detail(request, pk): #page displaying details of the member to be discharged.
    member = get_object_or_404(Member_Detail, pk=pk)
    admission_details = Admission_details.objects.filter(member=member).first()

     #calculating the number of days since admission
    days_since_admission = None
    if admission_details and admission_details.admission_date:
        current_date = now().date() 
        admission_date = admission_details.admission_date
        days_since_admission = (current_date - admission_date).days

    context = {
        'member': member,
        'admission_details': admission_details,
        'days_since_admission': days_since_admission,
        'current_date': timezone.now().date(),
        } #context variables

    return render(request, 'template/admissions/discharge-page.html', context)


@login_required
def add_update(request, pk):
    member = get_object_or_404(Member_Detail, pk=pk)
    admission_details = Admission_details.objects.filter(member=member).first()
    update_form = UpdateForm(request.POST or None)

    if request.method == 'POST':
        if 'add_update' in request.POST:
            if update_form.is_valid():
                update = update_form.save(commit=False)
                update.admission = admission_details
                update.save()

                return redirect('discharging_member_detail', pk=pk)
    context = {
                    'member': member,
                    'admission_details': admission_details,
                    'update_form': update_form,
                    'current_date': timezone.now().date(),
                }
    return render(request, 'discharged/discharged_members.html', context)
@login_required
def discharge_member(request, pk):
    member = get_object_or_404(Member_Detail, pk=pk)
    admission_details = Admission_details.objects.filter(member=member).first()
    
    if request.method == 'POST':
        discharge_details = Discharge_details(
            provider=admission_details.Provider,
            admission_date=admission_details.admission_date,
            admission_diagnosis=admission_details.admission_diagnosis,
            cover_used=admission_details.cover_used,
            initial_cover_value=admission_details.initial_cover_value,
            initial_cover_balance=admission_details.initial_cover_balance,
            requested_amount=admission_details.requested_amount,
            lou_issued=admission_details.lou_issued,
            admitted_by=admission_details.admited_by,
            relationship=member.relationship,
            name=member.name,
            membership_number=member.membership_number,
            payer=member.payer,
            scheme=member.scheme,
            status=member.status,
            validity=member.validity,
            discharge_date=request.POST.get('discharge_date'),
            final_approved_amount=request.POST.get('final_approved_amount', ''),
            discharge_notes=request.POST.get('discharge_notes', ''),
            discharged_by=f"{request.user.first_name} {request.user.last_name}"
        )
        discharge_details.save()

        context = {
            'member': member,
            'admission_details': admission_details,
            'discharge_details': discharge_details,
            'updates': Daily_update.objects.filter(admission=admission_details).order_by('-date'),
            'current_date': timezone.now().date(),
        }

        html_string = render_to_string('template/admissions/discharge_lou.html', context)
        pdf_io = io.BytesIO()
        HTML(string=html_string).write_pdf(target=pdf_io)
        pdf_io.seek(0)

        response = FileResponse(pdf_io, as_attachment=True, filename=f'Discharge_LOU_of-{member.name}.pdf')
        
        member.delete()

        return response

    updates = Daily_update.objects.filter(admission=admission_details).order_by('-date')

    context = {
        'member': member,
        'admission_details': admission_details,
        'updates': updates,
        'current_date': timezone.now().date(),
    }
    return render(request, 'discharged/discharged_members.html', context)
    
@login_required
def discharged_members(request): 
    query = request.GET.get('q')
    payer_filter = request.GET.get('payer')
    scheme_filter = request.GET.get('scheme')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    order = request.GET.get('order', '-discharge_date')  # Default order field

    discharged_entries = Discharge_details.objects.all()  # Get all discharge entries

    # Filtering based on search query
    if query:
        discharged_entries = discharged_entries.filter(
            Q(name__icontains=query) |
            Q(membership_number__icontains=query)
        )
    
    # Filtering based on payer
    if payer_filter:
        discharged_entries = discharged_entries.filter(payer__icontains=payer_filter)
    
    # Filtering based on scheme
    if scheme_filter:
        discharged_entries = discharged_entries.filter(scheme__icontains=scheme_filter)
    
    # Filtering based on discharge date
    if date_from:
        try:
            date_from = parse_date(date_from)
            discharged_entries = discharged_entries.filter(discharge_date__gte=date_from)
        except ValueError:
            # Handle invalid date format
            pass
    
    if date_to:
        try:
            date_to = parse_date(date_to)
            discharged_entries = discharged_entries.filter(discharge_date__lte=date_to)
        except ValueError:
            # Handle invalid date format
            pass

    # Ensure ordering field is valid
    valid_order_fields = ['discharge_date', 'membership_number', 'name', 'payer', 'scheme']
    if order not in valid_order_fields:
        order = '-discharge_date'

    discharged_entries = discharged_entries.order_by(order)

    context = {
        'discharged_entries': discharged_entries,
        'query': query,
        'payer_filter': payer_filter,
        'scheme_filter': scheme_filter,
        'date_from': date_from,
        'date_to': date_to,
        'order': order,
    }

    return render(request, 'template/admissions/discharged.html', context)


@login_required
def admission_history(request, discharge_id):
    # Get the discharge details based on the ID
    discharge_entry = get_object_or_404(
        Discharge_details, pk=discharge_id
    )

    # Fetch the member's previous admissions using Discharge_details model
    previous_admissions = Discharge_details.objects.filter(
        name=discharge_entry.name,
        membership_number=discharge_entry.membership_number
        ).order_by('-discharge_date')

    context = {
       'discharge_entry': discharge_entry,  # The specific entry being viewed
        'previous_admissions': previous_admissions,  # All entries with the same name
    }
    
    return render(request, 'template/admissions/discharge-history-page.html', context)

 #**************************************************************************************************************************************************
@login_required
def trend_analysis(request):
    # Fetch query parameters for filtering
    scheme_filter = request.GET.get('scheme')
    payer_filter = request.GET.get('payer')
    provider_filter = request.GET.get('provider')

    # Fetch admission and discharge data
    admissions = Admission_details.objects.all()
    discharges = Discharge_details.objects.all()

    # Apply filters if any
    if scheme_filter:
        admissions = admissions.filter(member__scheme=scheme_filter)
        discharges = discharges.filter(scheme=scheme_filter)
    if payer_filter:
        admissions = admissions.filter(member__payer=payer_filter)
        discharges = discharges.filter(payer=payer_filter)
    if provider_filter:
        admissions = admissions.filter(provider__name=provider_filter)
        discharges = discharges.filter(provider=provider_filter)

    # Group by date and count the admissions and discharges
    admission_trend = (
        admissions.values('admission_date')
        .annotate(count=Count('id'))
        .order_by('admission_date')
    )

    discharge_trend = (
        discharges.values('discharge_date')
        .annotate(count=Count('id'))  # Count discharges as independent units
        .order_by('discharge_date')
    )

    # Convert to DataFrame
    admission_df = pd.DataFrame(list(admission_trend))
    discharge_df = pd.DataFrame(list(discharge_trend))

    # Ensure DataFrames are not empty and handle the renaming accordingly
    if not admission_df.empty:
        admission_df.columns = ['date', 'admissions']
    else:
        admission_df = pd.DataFrame(columns=['date', 'admissions'])

    if not discharge_df.empty:
        discharge_df.columns = ['date', 'discharges']
    else:
        discharge_df = pd.DataFrame(columns=['date', 'discharges'])

    # Create a complete date range from the earliest to the latest date
    if not admission_df.empty and not discharge_df.empty:
        min_date = min(admission_df['date'].min(), discharge_df['date'].min())
        max_date = max(admission_df['date'].max(), discharge_df['date'].max())
    elif not admission_df.empty:
        min_date = admission_df['date'].min()
        max_date = admission_df['date'].max()
    elif not discharge_df.empty:
        min_date = discharge_df['date'].min()
        max_date = discharge_df['date'].max()
    else:
        min_date = datetime.now().date()
        max_date = datetime.now().date()

    complete_date_range = pd.date_range(start=min_date, end=max_date)

    # Ensure all dates are covered
    admission_df = admission_df.set_index('date').reindex(complete_date_range, fill_value=0).reset_index()
    discharge_df = discharge_df.set_index('date').reindex(complete_date_range, fill_value=0).reset_index()

    # Rename the 'index' column back to 'date'
    admission_df.columns = ['date', 'admissions']
    discharge_df.columns = ['date', 'discharges']

    # Merge DataFrames on the date
    trend_df = pd.merge(admission_df, discharge_df, on='date', how='outer').fillna(0)

    # Create the line graph using Plotly
    fig = px.line(trend_df, x='date', y=['admissions', 'discharges'],
                  labels={'value': 'Number of Members', 'date': 'Date'},
                  title='Trend of Admissions and Discharges')

    # Convert the plot to HTML
    graph_html = fig.to_html(full_html=False)

    context = {
        'graph': graph_html,
        'scheme_filter': scheme_filter,
        'payer_filter': payer_filter,
        'provider_filter': provider_filter,
    }

    return render(request, 'analytics/trend_analysis.html', context)


#***************************************************************************
# Reporting

@login_required
def reports(request):
    admitted = Admission_details.objects.filter(member__admission_status='admitted')
    discharged = Discharge_details.objects.all()

    total_admitted_cases = admitted.count()
    total_admitted_lou = Admission_details.objects.aggregate(Sum('lou_issued'))['lou_issued__sum'] or 0
    total_discharged_cases = discharged.count()

    
    admitted_payer_annotate= Admission_details.objects.annotate(
        payer_new = (F('member__payer'))
    )

    discharges_payer_annotate = Discharge_details.objects.annotate(
        payer_new = F('payer')
    )
    
    admitted_payers = admitted_payer_annotate.values('payer_new').distinct()
    discharged_payers = (discharges_payer_annotate.values('payer_new').distinct())
    
    payers = admitted_payers.union (discharged_payers)

    payer_data = []

    total_ip_cases = total_admitted_cases + total_discharged_cases

    for payer in payers:
        payer_name= payer.get('payer_new')

        total_admitted_payer = Admission_details.objects.filter(member__payer=payer_name, member__admission_status='admitted').count()
        total_admitted_lou_payer = Admission_details.objects.filter(member__payer=payer_name, member__admission_status='admitted').aggregate(Sum('lou_issued'))['lou_issued__sum'] or 0
        total_discharged_payer = Discharge_details.objects.filter(payer=payer_name).count()
        total_payer_cases = total_admitted_payer + total_discharged_payer

        payer_data.append({
            'payer': payer_name,
            'total_admitted_payer': total_admitted_payer,
            'total_admitted_lou_payer': total_admitted_lou_payer,
            'total_discharged_payer': total_discharged_payer,
            'total_payer_cases': total_payer_cases
        })

    context = {
        'total_admitted_cases': total_admitted_cases,
        'total_discharged_cases': total_discharged_cases,
        'total_ip_cases': total_ip_cases,
        'total_admitted_lou': total_admitted_lou,
        'payer_data': payer_data
    }
    return render(request, 'template/report/reports.html' , context)

@login_required
def payer_reports(request, payer_name): 
    admitted_members = Admission_details.objects.filter(member__payer=payer_name, member__admission_status='admitted')
    discharged_members = Discharge_details.objects.filter(payer=payer_name)

    #payer-specific statistics
    total_admitted_payer = admitted_members.count()
    total_admitted_lou_payer = admitted_members.aggregate(Sum('lou_issued'))['lou_issued__sum'] or 0
    total_discharged_payer = discharged_members.count()
    total_discharged_lou_payer = discharged_members.aggregate(Sum('final_approved_amount'))['final_approved_amount__sum'] or 0
    total_payer_cases = total_admitted_payer + total_discharged_payer
    all_inpatient_cases = list(admitted_members) + list(discharged_members)

    current_date = now().date()
    for member in admitted_members:
        admission_date = member.admission_date
        los = (current_date - admission_date).days
        member.los = los #(temporarily saves this to the object)

    context = {
        'payer_name': payer_name,
        'total_admitted_payer': total_admitted_payer,
        'total_admitted_lou_payer': total_admitted_lou_payer,
        'total_discharged_payer': total_discharged_payer,
        'total_payer_cases': total_payer_cases,
        'admitted_members': admitted_members,
        'discharged_members': discharged_members,
        'total_discharged_lou_payer': total_discharged_lou_payer,
        'all_inpatient_cases': all_inpatient_cases,
    }
    return render(request, 'template/report/reports-view.html', context)

def export_payer_report(request, payer_name):
    admitted_members= Admission_details.objects.filter(member__payer=payer_name, member__admission_status='admitted')
    discharged_members= Discharge_details.objects.filter(payer=payer_name)

    current_date= now().date()
    for member in admitted_members:
        admission_date = member.admission_date
        los= (current_date - admission_date).days
        member.los = los

    for update in admitted_members:
        last_update = update.updates.order_by('-date').first()
        interim_bill = last_update.interim_bill if last_update else None
        date_of_interim_bill = last_update.date if last_update else None 
        if date_of_interim_bill is not None:
            if not is_naive(date_of_interim_bill):  # Check if the datetime is naive
                date_of_interim_bill = make_naive(date_of_interim_bill)
        else:
            date_of_interim_bill= " "

    wb = openpyxl.Workbook()
    ws= wb.active
    ws.title = f"{payer_name}'s Report"

    headers= [
        "Beneficiary", "Membership No.", "scheme", "Relationship",
        "Date of Admission", "Diagnosis", "Admitting Provider", "Ammount Requested", 
        "Initial LOU issued", "Cover Benefits Applied", "Available Cover Balance",
        "Interim Bill", "Date of Interim Bill", 
        "LOS", "Date of Discharge", "Final Bill", "Final LOU Issued", "Status",
        "Admited By", "Discharged By",
    ]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    row_num = 2
    for member in admitted_members:
        ws[f"A{row_num}"] = member.member.name
        ws[f"B{row_num}"] = member.member.membership_number
        ws[f"C{row_num}"] = member.member.scheme
        ws[f"D{row_num}"] = member.member.relationship
        ws[f"E{row_num}"] = member.admission_date
        ws[f"F{row_num}"] = member.admission_diagnosis
        ws[f"G{row_num}"] = member.Provider.name
        ws[f"H{row_num}"] = member.requested_amount
        ws[f"I{row_num}"] = member.lou_issued
        ws[f"J{row_num}"] = member.cover_used
        ws[f"K{row_num}"] = member.initial_cover_balance
        ws[f"L{row_num}"] = interim_bill
        ws[f"M{row_num}"] = date_of_interim_bill
        ws[f"N{row_num}"] = member.los 
        ws[f"O{row_num}"] = "Admitted"
        ws[f"P{row_num}"] = "Admitted" #final bill
        ws[f"Q{row_num}"] = "-" #FInal lou
        ws[f"R{row_num}"] = "-" #exgratia
        ws[f"S{row_num}"] = member.admited_by
        ws[f"T{row_num}"] = "Admitted"
       
        ws[f"H{row_num}"].number_format = '#,##0.00' 
        ws[f"I{row_num}"].number_format = '#,##0.00'
        row_num += 1

    for member in discharged_members:
        ws[f"A{row_num}"] = member.name
        ws[f"B{row_num}"] = member.membership_number
        ws[f"C{row_num}"] = member.scheme
        ws[f"D{row_num}"] = member.relationship
        ws[f"E{row_num}"] = member.admission_date
        ws[f"F{row_num}"] = member.admission_diagnosis
        ws[f"G{row_num}"] = member.provider
        ws[f"H{row_num}"] = member.requested_amount
        ws[f"I{row_num}"] = member.lou_issued
        ws[f"J{row_num}"] = member.cover_used
        ws[f"K{row_num}"] = member.initial_cover_balance
        ws[f"L{row_num}"] = "Discharged"
        ws[f"M{row_num}"] = "Discharged"
        ws[f"N{row_num}"] = member.days_admitted  
        ws[f"O{row_num}"] = member.discharge_date
        ws[f"P{row_num}"] = member.discharge_notes
        ws[f"Q{row_num}"] = member.final_approved_amount
        ws[f"R{row_num}"] = "Discharged"
        ws[f"S{row_num}"] = member.admitted_by
        ws[f"T{row_num}"] = member.discharged_by

        ws[f"H{row_num}"].number_format = '#,##0.00' 
        ws[f"I{row_num}"].number_format = '#,##0.00'
        ws[f"P{row_num}"].number_format = '#,##0.00' 
        row_num += 1

    summary_ws = wb.create_sheet(title=f"{payer_name}'s Summary")
    summary_ws.append(["Scheme", "Total LOU Issued", "Number of Admissions", "Distribution"])

    summary_data = admitted_members.values('member__scheme').annotate(
        total_lou=Sum('lou_issued'),
        admitted_members_count=Count('id')
    ).filter(admitted_members_count__gt=0)

    total_active_cases = sum(item['admitted_members_count'] for item in summary_data)
    distribution = (admitted_members.count() / total_active_cases *100)if total_active_cases > 0 else 0

    for summary in summary_data:
        summary_ws.append([
            summary.get('member__scheme', 'N/A'),
            summary.get('total_lou', 0),
            summary.get('admitted_members_count', 0),
            f"{distribution:.2f}%"
        ])



    for sheet in [ws, summary_ws]:
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) 
                                for cell in col if cell.value) + 2
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{payer_name}_report_{current_date}.xlsx"'        

    wb.save(response)
    return response

#***************************************************************************
#Accounts management

@login_required
def user_account(request):
    """View to display user account details."""
    return render(request, 'template/user/user-page.html')

@login_required
def update_user_details(request):
    """View to handle updating user personal details."""
    if request.method == 'POST':
        form = UserDetailForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your details have been updated successfully.')
            return redirect('user_account')
    else:
        form = UserDetailForm(instance=request.user)
    
    return render(request, 'user_account/update_user_details.html', {'form': form})

@login_required
def change_password(request):
    """View to handle password change."""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important for keeping the user logged in
            messages.success(request, 'Your password was successfully updated!')
            return redirect('user_account')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'user_account/change_password.html', {'form': form})

#***************************************************************************
#Schemes

@login_required
def schemes(request): #list of schemes.
    query = request.GET.get('q')
    payer_filter = request.GET.get('payer')

    schemes = Scheme.objects.all()

    if query:
        schemes = schemes.filter(
            Q(name__icontains=query)
        )
    if payer_filter:
        schemes = schemes.filter(payer__icontains=payer_filter)
    
    return render(request, 'template/schemes/scheme-list.html', 
                  {'schemes': schemes, 
                    'query': query, 
                    'payer_filter': payer_filter})

from .forms import SchemeAdminForm

@login_required
def scheme_detail(request, scheme_id):
    scheme = get_object_or_404(Scheme, id=scheme_id)
    if request.method == 'POST':
        form = SchemeAdminForm(request.POST, instance=scheme)
        if form.is_valid():
            scheme = form.save(commit=False)
            selected_providers = form.cleaned_data.get('service_provider')
            messages.success(request, 'Scheme details updated successfully.')
            scheme.service_provider.add(*selected_providers)
            messages.success(request, 'Scheme details updated successfully.')
            return redirect('scheme_detail', scheme_id=scheme_id)
        else:
            print(form.errors)
    else:
        form = SchemeAdminForm(instance=scheme)

    return render(request, 'template/schemes/schemes.html', {
        'scheme': scheme,
        'form': form,
    })

def manage_schemes_providers(request):
    schemes = Scheme.objects.all()
    providers = Provider.objects.all()
    
    scheme_form = SchemeForm()
    provider_form = ProviderForm()
    
    if request.method == 'POST':
        if 'upload_schemes' in request.POST:
            scheme_form = SchemeForm(request.POST, request.FILES)
            if scheme_form.is_valid():
                schemes_csv = request.FILES['schemes_csv']
                decoded_file = schemes_csv.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded_file)
                next(reader)  # Skip the header row

                for row in reader:
                    scheme_name = row[0]
                    Scheme.objects.get_or_create(name=scheme_name.strip())
                
                return redirect(reverse('success_url'))  # Replace 'success_url' with the actual named URL
        elif 'upload_providers' in request.POST:
            provider_form = ProviderForm(request.POST, request.FILES)
            if provider_form.is_valid():
                providers_csv = request.FILES['providers_csv']
                decoded_file = providers_csv.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded_file)
                next(reader)  # Skip the header row

                for row in reader:
                    provider_name, scheme_name = row
                    provider, _ = Provider.objects.get_or_create(name=provider_name.strip())
                    scheme = Scheme.objects.filter(name=scheme_name.strip()).first()
                    if scheme:
                        scheme.providers.add(provider)
                
                return redirect(reverse('success_url'))  # Replace 'success_url' with the actual named URL

    return render(request, 'manage/manage_schemes_providers.html', {
        'schemes': schemes,
        'providers': providers,
        'scheme_form': scheme_form,
        'provider_form': provider_form
    })

def edit_scheme(request, scheme_id):
    scheme = get_object_or_404(Scheme, id=scheme_id)
    # Handle editing the scheme
    return render(request, 'manage/edit_scheme.html', {'scheme': scheme})

def delete_scheme(request, scheme_id):
    scheme = get_object_or_404(Scheme, id=scheme_id)
    scheme.delete()
    return redirect(reverse('manage/manage_schemes_providers'))

#***************************************************************************
#providers
@login_required
def providers(request):
    query = request.GET.get('q')
    provider_filter = request.GET.get('provider')

    providers = Provider.objects.all()
    if query:
        providers= providers.filter(
            Q(name__icontains=query))
    if provider_filter:
        providers = providers.filter(
            Q(name__icontains=provider_filter))
        
    return render(request, 'template/providers/provider-list.html', 
                  {'providers': providers, 
                    'query': query, 
                    'provider_filter': provider_filter})

@login_required
def provider_detail(request, provider_id):
    provider = get_object_or_404(Provider, id=provider_id)
    return render(request, 'template/providers/providers.html', {'provider': provider})

def view_providers(request, scheme_id):
    scheme = get_object_or_404(Scheme, id=scheme_id)
    providers = scheme.providers.all()
    return render(request, 'manage/view_providers.html', {'scheme': scheme, 'providers': providers})

def edit_provider(request, provider_id):
    provider = get_object_or_404(Provider, id=provider_id)
    # Handle editing the provider
    return render(request, 'manage/edit_provider.html', {'provider': provider})

def delete_provider(request, provider_id):
    provider = get_object_or_404(Provider, id=provider_id)
    provider.delete()
    return redirect(reverse('manage/manage_schemes_providers'))

def view_schemes(request, provider_id):
    provider = get_object_or_404(Provider, id=provider_id)
    schemes = provider.schemes.all()
    return render(request, 'manage/view_schemes.html', {'provider': provider, 'schemes': schemes})

def add_scheme(request):
    if request.method == 'POST':
        # Check if the bulk input field is used
        bulk_schemes = request.POST.get('bulk_schemes', '')
        if bulk_schemes:
            # Handle bulk entry
            scheme_names = bulk_schemes.split('\n')
            for name in scheme_names:
                name = name.strip()
                if name:  # Ensure not adding empty entries
                    Scheme.objects.create(name=name)
            return redirect('manage_schemes_providers')

        # Handle single entry using form
        form = SchemeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('manage_schemes_providers')
    else:
        form = SchemeForm()
    
    return render(request, 'manage/manage_schemes_providers.html', {'form': form})


